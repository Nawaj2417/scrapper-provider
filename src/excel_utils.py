from openpyxl import load_workbook
import asyncio
from src.scraper_async import scrape_website

BATCH_SIZE = 10

def chunked(lst, size):
    for i in range(0, len(lst), size):
        yield lst[i:i + size]

async def process_excel(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    required = ["email", "facebook", "instagram", "scrape_status"]
    for col in required:
        if col not in headers:
            ws.cell(row=1, column=len(headers)+1, value=col)
            headers.append(col)
    h = {name: idx+1 for idx, name in enumerate(headers)}

    # Collect rows + websites
    rows = []
    websites = []

    for row in range(2, ws.max_row + 1):
        website = ws.cell(row=row, column=h["Website"]).value
        rows.append(row)
        websites.append(website)

    # Process in batches
    for batch_rows, batch_sites in zip(
        chunked(rows, BATCH_SIZE),
        chunked(websites, BATCH_SIZE)
    ):
        tasks = [scrape_website(site) for site in batch_sites]
        results = await asyncio.gather(*tasks)

        for row, result in zip(batch_rows, results):
            if not result:
                ws.cell(row=row, column=h["scrape_status"], value="invalid_url")
                continue

            if result["emails"]:
                ws.cell(row=row, column=h["email"], value=result["emails"][0])
            if result["facebook"]:
                ws.cell(row=row, column=h["facebook"], value=result["facebook"])
            if result["instagram"]:
                ws.cell(row=row, column=h["instagram"], value=result["instagram"])

            ws.cell(
                row=row,
                column=h["scrape_status"],
                value="success" if any(result.values()) else "no_contact"
            )

        # 🔒 SAVE after each batch
        wb.save(output_path)

    wb.save(output_path)
